#!/usr/bin/env python
'''Use the questionnaire client to generate a Excel spreadsheet
This is intended to serve as an example; so rather than loading all the selections for all proposals, I pick a small subset
These are the columns that are added.
Proposal ID (proposal_id)
Instrument (instrument)
XRay Mode (xray-mode)
XRay Energy (xray-energy-1)
XRay Techniques to be used (xraytech-tech-1)
'''

import argparse
from openpyxl import Workbook

from psdm_qs_cli import QuestionnaireClient


def generateExcelSpreadSheetForRun(qs, run, excelFilePath):
    '''
    Generate a Excel spreadsheet with data from a run.
    :param: qs - A Questionnaire client
    :run: The number number/name; this is a string like run15 which is what the questionnaire uses in its URL
    '''

    nameMappings = qs.formLabelMappings(run)
    column2Names = [('proposal_id', "Proposal")]
    column2Names.extend([(n,v) for n, v in nameMappings.items()])

    wb = Workbook()
    ws = wb.active
    ws.title = run

    # Generate the header column using the 2 element in the column2Names tuples
    for clnum in range(len(column2Names)):
        _ = ws.cell(row=1, column=clnum+1, value=column2Names[clnum][1])


    # Get a list of proposals
    proposals = qs.getProposalsListForRun(run)
    rowNum = 1
    for proposalid in sorted(proposals.keys()):
        print("Getting details for proposal ", proposalid)
        proposalDetails = qs.getProposalDetailsForRun(run, proposalid)
        # Add the details of each proposal to the information obtained from the proposal list call.
        proposals[proposalid].update(proposalDetails)
        for clnum in range(len(column2Names)):
            ckey = column2Names[clnum][0]
            if ckey in proposals[proposalid]:
                _ = ws.cell(row=rowNum+1, column=clnum+1, value=proposals[proposalid][ckey])
            else:
                _ = ws.cell(row=rowNum+1, column=clnum+1, value='')
        rowNum = rowNum + 1

    wb.save(excelFilePath)
    print("Saved data into", excelFilePath)


def main():
    parser = argparse.ArgumentParser(description='Load data from the questionnaire into a an Excel spreadsheet')
    parser.add_argument('--questionnaire_url', default="https://pswww.slac.stanford.edu/ws-kerb/questionnaire")
    parser.add_argument('--no_kerberos', action="store_false")
    parser.add_argument('run')
    parser.add_argument('excelFilePath')
    args = parser.parse_args()

    qs = QuestionnaireClient(args.questionnaire_url, args.no_kerberos)
    generateExcelSpreadSheetForRun(qs, args.run, args.excelFilePath)


if __name__ == '__main__':
    main()
